import tkinter as tk
from tkinter import ttk, PhotoImage, filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
import matplotlib.pyplot as plt
from text_analysis import TextAnalyzer
from menu import MenuBar
from importar import leer_pdf, leer_txt, leer_docx
from analisis_avanzado import abrir_analisis_avanzado
from chart_converter import convertir_a_grafica
from audio import transcribe_audio
import time
from threading import Thread

class Aplicacion:
    def __init__(self, root):
        self.root = root
        self.root.title("BrainLingua NLP")
        self.root.state('zoomed')  
        self.root.configure(bg="#2e2e2e")
        
        self.analisis_realizado = False
        
        self.menu_bar = MenuBar(root)
        self._setup_styles()
        self._setup_widgets()

        self.analyzer = TextAnalyzer()
        self.stored_text = ""

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", 
                        background="#2e2e2e",
                        foreground="white", 
                        fieldbackground="#2e2e2e",
                        rowheight=25)
        style.configure("Treeview.Heading",
                        background="#404040",
                        foreground="white")
        style.configure("TButton", 
                        background="#207567",  
                        foreground="white", 
                        padding=10)
        style.map("TButton", 
                background=[('active', '#1a5c52'), ('pressed', '#207567')])
        style.configure("TLabel",
                        background="#2e2e2e",
                        foreground="white")
        style.configure("TEntry",
                        background="#2e2e2e",
                        foreground="white",
                        fieldbackground="#2e2e2e")

    def _setup_widgets(self):
        # Agregar logotipo y texto de bienvenida
        self._add_logo()
        self._add_welcome_text()

        # Crear frame principal
        self.main_frame = tk.Frame(self.root, bg="#2e2e2e")
        self.main_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

        # Crear frame izquierdo para los botones
        self.left_frame = tk.Frame(self.main_frame, bg="#2e2e2e")
        self.left_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # Crear frame derecho para el cuadro de texto y la tabla de resultados
        self.right_frame = tk.Frame(self.main_frame, bg="#2e2e2e")
        self.right_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        # Agregar botones, cuadro de texto y tabla de resultados
        self._add_buttons(self.left_frame)
        self._add_text_box(self.right_frame)
        self._setup_treeview(self.right_frame)

        # Configurar pesos para las columnas y filas del grid
        self.main_frame.grid_columnconfigure(0, weight=1, minsize=200)
        self.main_frame.grid_columnconfigure(1, weight=4)
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(1, weight=1)

    def _add_logo(self):
        logo_image = Image.open("./src/img/prueba_logo.png").resize((40, 40))
        logo_photo = ImageTk.PhotoImage(logo_image)
        tk.Label(self.root, image=logo_photo, bg="#2e2e2e").grid(row=0, column=0, pady=(5, 0), padx=5, sticky="w")
        self.root.image = logo_photo  # Prevent garbage collection

    def _add_welcome_text(self):
        welcome_text = (
            "¡Bienvenido a BrainLingua!\n"
            "Este programa realiza análisis de texto por medio del lenguaje natural.\n"
            "Autor: Pablo Fernández Planas\n"
        )
        tk.Label(self.root, text=welcome_text, justify="left", bg="#2e2e2e", fg="white", font=("Helvetica", 12)).grid(
            row=0, column=1, columnspan=3, pady=(5, 0), padx=10, sticky="w"
        )

    def _add_text_box(self, parent_frame):
        text_frame = tk.Frame(parent_frame, bg="#2e2e2e")
        text_frame.grid(row=0, column=0, sticky="nsew")

        self.text_box = tk.Text(text_frame, width=80, height=15, borderwidth=2, relief="solid", bg='white', fg='black', insertbackground='black', highlightthickness=2, highlightbackground='#207567')
        self.text_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Agregar botón para borrar el texto
        imagen_boton_delete = PhotoImage(file="./src/img/Delete.png").subsample(8, 8)
        ttk.Button(text_frame, image=imagen_boton_delete, command=self.clear_text_box).pack(side=tk.RIGHT, padx=5, pady=5)
        self.root.image_delete = imagen_boton_delete

        # Agregar botón para realizar el análisis
        img_analysis = PhotoImage(file="./src/img/EJECUTAR.png").subsample(15, 15)
        ttk.Button(text_frame, image=img_analysis, text='Ejecutar', compound=tk.LEFT, command=self.store_and_display_analysis).pack(side=tk.RIGHT, padx=5, pady=5)
        self.root.image_analysis = img_analysis

        text_frame.grid_rowconfigure(0, weight=1)
        text_frame.grid_columnconfigure(0, weight=1)

    def _add_buttons(self, parent_frame):
        button_options = {"width": 40}

        # Cargar las imágenes para los botones
        img_import_pdf = PhotoImage(file="./src/img/PDF.png").subsample(10, 10)
        img_import_txt = PhotoImage(file="./src/img/TXT.png").subsample(10, 10)
        img_import_docx = PhotoImage(file="./src/img/DOCX.png").subsample(10, 10)
        img_advanced_analysis = PhotoImage(file="./src/img/ANALISIS_AVANZADO.png").subsample(15, 15)
        img_transcribe_audio = PhotoImage(file="./src/img/MP3.png").subsample(10, 10)
        img_convert_graph = PhotoImage(file="./src/img/GRAFICA.png").subsample(10, 10)

        # Almacenar las imágenes para evitar que se recojan como basura
        self.root.images = {
            "import_pdf": img_import_pdf,
            "import_txt": img_import_txt,
            "import_docx": img_import_docx,
            "advanced_analysis": img_advanced_analysis,
            "transcribe_audio": img_transcribe_audio,
            "convert_graph": img_convert_graph
        }

        # Definir los botones y sus comandos
        buttons = [
            ("Importar PDF", lambda: leer_pdf(self.text_box), img_import_pdf),
            ("Importar TXT", lambda: leer_txt(self.text_box), img_import_txt),
            ("Importar DOCX", lambda: leer_docx(self.text_box), img_import_docx),
            ("Análisis avanzado", lambda: abrir_analisis_avanzado(self.root, self.text_box.get("1.0", tk.END)), img_advanced_analysis),
            ("Importar Audio", self.import_audio, img_transcribe_audio),
            ("Convertir a gráfica", lambda: convertir_a_grafica(self), img_convert_graph)
        ]

        # Agregar los botones al framee
        for i, (text, command, image) in enumerate(buttons):
            ttk.Button(parent_frame, text=text, command=command, image=image, compound=tk.LEFT, **button_options).pack(pady=5)

    def _setup_treeview(self, parent_frame):
        # Crear el Treeview para mostrar los resultados del análisis
        self.tree = ttk.Treeview(parent_frame, show="headings")
        self.tree.grid(row=1, column=0, sticky="nsew")  # Treeview en la fila 1
        parent_frame.grid_rowconfigure(1, weight=1)
        parent_frame.grid_columnconfigure(0, weight=1)

        # Ajustar las columnas para que se autoescale al contenido
        self.tree.column("#0", stretch=tk.YES)
        for col in self.tree["columns"]:
            self.tree.column(col, stretch=tk.YES)

        # Crear un frame para el botón debajo del Treeview
        button_frame = tk.Frame(parent_frame, bg="#2e2e2e")
        button_frame.grid(row=2, column=0, sticky="ew")  # Botón en la fila 2

        # BOTON SAVE
        img_save = PhotoImage(file="./src/img/SAVE.png").subsample(15, 15)
        ttk.Button(button_frame, image=img_save, text='Guardar', compound=tk.LEFT, command=self.exportar_a_excel).pack(side=tk.RIGHT, padx=5, pady=5)
        self.root.img_export_excel = img_save

        parent_frame.grid_rowconfigure(2, weight=0)
        parent_frame.grid_columnconfigure(0, weight=1)

    def store_and_display_analysis(self):
        # Limpiar resultados anteriores
        self.tree.delete(*self.tree.get_children())
        self.stored_text = self.text_box.get("1.0", tk.END).strip()
        if not self.stored_text:
            return
        self.analisis_realizado = True

        # Analizar el texto y obtener las estadísticas
        analysis_results = self.analyzer.analyze_text(self.stored_text)

        # Definir las columnas de la tabla si hay resultados
        if analysis_results:
            columns = list(analysis_results[0].keys())
            self.tree["columns"] = columns

            # Configurar las cabeceras de las columnas
            for col in columns:
                self.tree.heading(col, text=col)
                self.tree.column(col, anchor=tk.CENTER)

            # Insertar los valores en la tabla
            for result in analysis_results:
                values = [result[col] for col in columns]
                self.tree.insert("", "end", values=values)

            # Ajustar las columnas para que se autoescalen al contenido
            for col in self.tree["columns"]:
                self.tree.column(col, stretch=tk.YES)

    def clear_text_box(self):
        # Limpiar el cuadro de texto
        self.text_box.delete("1.0", tk.END)

    def exportar_a_excel(self):
        if not self.analisis_realizado:
            messagebox.showwarning("Exportar a Excel", "Por favor, realice un análisis antes de exportar los datos.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])
        if not file_path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        for idx, col in enumerate(self.tree["columns"], start=1):
            sheet.cell(row=1, column=idx, value=col)

        for idx, item in enumerate(self.tree.get_children(), start=2):
            for col_idx, col in enumerate(self.tree["columns"], start=1):
                sheet.cell(row=idx, column=col_idx, value=self.tree.set(item, col))

        workbook.save(file_path)
        messagebox.showinfo("Exportar a Excel", "Los datos han sido exportados correctamente.")

    def import_audio(self):
        # Abrir cuadro de diálogo para seleccionar archivo de audio WAV
        audio_file_path = filedialog.askopenfilename(filetypes=[("Archivos de audio WAV", "*.wav")])
        if audio_file_path:
            # Crear la ventana emergente de espera
            self.wait_popup = tk.Toplevel(self.root)
            self.wait_popup.title("Procesando audio")
            
            # Obtener las dimensiones de la pantalla
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()

            # Calcular las coordenadas para centrar la ventana emergente en la pantalla
            x_coordinate = (screen_width - 300) // 2  # Ancho de la ventana emergente es 300
            y_coordinate = (screen_height - 100) // 2  # Alto de la ventana emergente es 100

            # Establecer la geometría de la ventana emergente para centrarla en la pantalla
            self.wait_popup.geometry(f"300x100+{x_coordinate}+{y_coordinate}")
            
            self.wait_popup.configure(bg="#2e2e2e")
            self.wait_popup.resizable(False, False)
            self.wait_popup.attributes('-topmost', 'true')  # Mantener la ventana arriba
            tk.Label(self.wait_popup, text="Espere, se está procesando su audio...", bg="#2e2e2e", fg="white").pack(pady=10)
            
            # Agregar la barra de progreso
            self.progress_bar = ttk.Progressbar(self.wait_popup, orient="horizontal", length=200, mode="indeterminate")
            self.progress_bar.pack(pady=10)

            # Iniciar la animación de la barra de progreso
            self.progress_bar.start()

            # Iniciar un hilo para procesar el audio mientras se muestra la ventana emergente
            thread = Thread(target=self.process_audio, args=(audio_file_path,))
            thread.start()

    def process_audio(self, audio_file_path):
        # Realizar la transcripción del audio
        transcription = transcribe_audio(audio_file_path)
        
        # Detener la animación de la barra de progreso
        self.progress_bar.stop()

        # Una vez completado el procesamiento, actualizar la barra de progreso y mostrar la transcripción en el cuadro de texto
        self.wait_popup.destroy()  # Cerrar la ventana emergente de espera
        self.text_box.insert(tk.END, transcription)


def main():
    # Iniciar la aplicación
    root = tk.Tk()
    app = Aplicacion(root)
    root.mainloop()

if __name__ == "__main__":
    main()

