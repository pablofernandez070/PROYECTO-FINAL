import openpyxl
from tkinter import filedialog, messagebox

def exportar_a_excel(tree, analisis_realizado):
    if not analisis_realizado:
        messagebox.showwarning("Exportar a Excel", "Por favor, realice un análisis antes de exportar los datos.")
        return

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])
    if not file_path:
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for idx, col in enumerate(tree["columns"], start=1):
        sheet.cell(row=1, column=idx, value=col)

    for idx, item in enumerate(tree.get_children(), start=2):
        for col_idx, col in enumerate(tree["columns"], start=1):
            sheet.cell(row=idx, column=col_idx, value=tree.set(item, col))

    workbook.save(file_path)
    messagebox.showinfo("Exportar a Excel", "Los datos han sido exportados correctamente.")